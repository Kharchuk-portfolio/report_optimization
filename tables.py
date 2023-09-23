import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# change every week
week = 38

# function for create new column
def categorize_views(view_count):
    if view_count > 100000:
        return 'Більше 100 тис'
    else:
        return 'До 100 тис'

# function for add percent
def add_percent(df, column_name, column_name_new):
    total_sum = df[column_name].sum()
    df[column_name_new] = df.apply(lambda row: 2*(row[column_name] / total_sum), axis=1)

# function for write a table to a file
def table_write(sheet, table):
    for r_idx, row in enumerate(dataframe_to_rows(table, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

common = pd.read_excel('База Ютюб дохід-.xlsx', sheet_name='Загалом')

common = common[common['Тиждень'] == week].iloc[:, :-5]
common['Перегляди група 2'] = common['Перегляди'].apply(categorize_views)

# pivot table (time download and type broadcast)
time_type = pd.pivot_table(common, values=['Назва відео', 'Приблизний дохід (USD)', 'Перегляди'],
                           index=['Час завантаження', 'Тип трансляції'],
                           aggfunc={'Назва відео': "count", 'Приблизний дохід (USD)': sum, 'Перегляди': sum},
                           margins=True, margins_name="Загальний підсумок")

time_type = time_type.rename(columns={'Назва відео': 'К-сть відео', 'Приблизний дохід (USD)': 'Дохід'})

add_percent(time_type, 'Дохід', 'Дохід, %')
add_percent(time_type, 'Перегляди', 'Перегляди, %')

time_type = time_type[['К-сть відео', 'Дохід', 'Дохід, %', 'Перегляди', 'Перегляди, %']]

# pivot table (time download and count views)
time_count = pd.pivot_table(common, values=['Назва відео','Приблизний дохід (USD)','Перегляди'], 
                            index=['Час завантаження', 'Перегляди група 2'], 
                            aggfunc={'Назва відео': "count", 'Приблизний дохід (USD)': sum, 'Перегляди': sum}, 
                            margins=True, margins_name="Загальний підсумок")

time_count = time_count.rename(columns={'Назва відео': 'К-сть відео', 'Приблизний дохід (USD)': 'Дохід'})

add_percent(time_count, 'Дохід', 'Дохід, %')
add_percent(time_count, 'Перегляди', 'Перегляди, %')

time_count = time_count[['К-сть відео', 'Дохід', 'Дохід, %', 'Перегляди', 'Перегляди, %']]

# open and write to a file
book = openpyxl.load_workbook("База Ютюб дохід-.xlsx")

sheet_1 = book["Час_зав+тип_транс"]
sheet_2 = book["Час_зав+перегляди"]

book.remove(sheet_1)
book.remove(sheet_2)

sheet1 = book.create_sheet()
sheet1.title = "Час_зав+тип_транс"
sheet2 = book.create_sheet()
sheet2.title = "Час_зав+перегляди"

table_write(sheet1, time_type)
table_write(sheet2, time_count)

book.save("База Ютюб дохід-.xlsx")

